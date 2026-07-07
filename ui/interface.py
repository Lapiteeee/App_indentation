import threading
import os
import subprocess
import tempfile

import flet as ft
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from flet.matplotlib_chart import MatplotlibChart

from config import CHEMINS_MAILLAGE
from core.io_donnees import charger_bdd_module, lire_courbe_fh
from core.fmax     import executer_optimisation_fmax,     ArretUtilisateur as _AU_fmax
from core.libre    import executer_optimisation_libre,    ArretUtilisateur as _AU_libre
from core.faible   import executer_optimisation_faible,   ArretUtilisateur as _AU_faible
from core.zero     import executer_optimisation_zero,     ArretUtilisateur as _AU_zero
from core.frelatif import executer_optimisation_frelatif, ArretUtilisateur as _AU_frelatif
from core.maths    import loi_hollomon

_ARRET_CLASSES = (_AU_fmax, _AU_libre, _AU_faible, _AU_zero, _AU_frelatif)

METHODES_STYLE = {
    'fmax':     ('tab:blue',   '-',  'o', 'Fmax'),
    'libre':    ('tab:orange', '--', 's', 'Pivot libre'),
    'faible':   ('tab:green',  ':',  '^', 'Pivot faible'),
    'zero':     ('tab:red',    '-.', 'D', 'Pivot zéro'),
    'frelatif': ('tab:purple', '-',  'P', 'F-relatif'),
}

_PIVOT_FIG_H = 4.5
_PIVOT_DPI   = 100
_PIVOT_H_PX  = int(_PIVOT_FIG_H * _PIVOT_DPI)   # 450
_RAIL_W      = 91   # NavigationRail + VerticalDivider
_PAD         = 40   # padding left+right in view_pivot


def main(page: ft.Page):
    page.title = "Identification de Loi d'Écrouissage"
    page.theme_mode       = ft.ThemeMode.LIGHT
    page.padding          = 0
    page.window_maximized = True

    etat = {"en_cours": False, "stop": False}

    IDX_CONFIG  = 0
    IDX_PIVOT   = 1
    IDX_RESULTS = 2

    # ── file picker ───────────────────────────────────────────────────────
    def _choisir_fichier(champ_texte):
        def _ouvrir():
            try:
                res = subprocess.run(
                    ["osascript", "-e", 'POSIX path of (choose file)'],
                    capture_output=True, text=True, timeout=60
                )
                chemin = res.stdout.strip()
                if chemin:
                    champ_texte.value = chemin
                    champ_texte.update()
            except Exception:
                pass
        threading.Thread(target=_ouvrir, daemon=True).start()

    # ── inputs ────────────────────────────────────────────────────────────
    input_E     = ft.TextField(label="Module d'Young E (MPa)", value="210000.0", width=200)
    input_sig_y = ft.TextField(label="Guess σ_y (MPa)",        value="860.0",   width=150)
    input_n     = ft.TextField(label="Guess n",                 value="0.36",   width=110)
    input_a     = ft.TextField(label="Rayon exp. a (µm) [opt]", value="0.0",   width=180)

    input_decalage = ft.TextField(
        label="Décalage initial (Pivot libre)",
        value="0.0", width=220, visible=False,
        tooltip="Décalage initial h_exp (µm). 0 pour courbe BDD.",
    )
    input_F_pivot = ft.TextField(
        label="F_pivot (N) — Pivot faible",
        value="10", width=200, visible=False,
        tooltip="Force (N) au point d'alignement.",
    )

    # ── checkboxes ────────────────────────────────────────────────────────
    cb_fmax     = ft.Checkbox(label="Fmax",         value=True)
    cb_h_max_manuel = ft.Checkbox(label="h_max manuel", value=False, visible=True)
    input_h_max = ft.TextField(
        label="h_max (µm)",
        value="", width=150, visible=False,
        tooltip="Valeur de h_max imposée (µm). Laissez vide pour la détection automatique.",
    )
    cb_libre    = ft.Checkbox(label="Pivot libre",  value=False)
    cb_faible   = ft.Checkbox(label="Pivot faible", value=False)
    cb_zero     = ft.Checkbox(label="Pivot zéro",   value=False)
    cb_frelatif = ft.Checkbox(label="F-relatif",    value=False)

    # ── pivot-zero state ──────────────────────────────────────────────────
    etat_pivot = {'h_pivot': None, 'h_exp': None, 'F_exp': None}

    fig_pivot, ax_pivot = plt.subplots(figsize=(14.0, _PIVOT_FIG_H))
    fig_pivot.subplots_adjust(left=0.07, right=0.98, bottom=0.14, top=0.88)
    ax_pivot.set_title("Chargez la courbe F-h puis cliquez pour définir l'origine",
                       color='gray', fontsize=10)
    ax_pivot.grid(True)
    chart_pivot = MatplotlibChart(fig_pivot, expand=True)

    txt_h_pivot = ft.Text(
        "Aucun point sélectionné — cliquez sur 'Charger la courbe' puis cliquez sur le graphe",
        italic=True, color=ft.Colors.GREY_600, size=11,
    )

    def _charger_courbe_pivot():
        chemin = txt_courbe.value
        if "Aucun" in chemin:
            txt_h_pivot.value = "⚠ Chargez d'abord la courbe F-h (vue Configuration)"
            txt_h_pivot.color = ft.Colors.ORANGE_700
            txt_h_pivot.update()
            return
        try:
            h_arr, F_arr = lire_courbe_fh(chemin)
            etat_pivot['h_exp'] = np.array(h_arr)
            etat_pivot['F_exp'] = np.array(F_arr)
            etat_zoom['xlim'] = None
            # Y min légèrement négatif pour que le 0 soit lisible avec espace en dessous
            F_np = np.array(F_arr)
            f_max = float(F_np.max())
            etat_zoom['ylim'] = (-f_max * 0.20, f_max * 1.08)
            _dessiner_pivot(h_sel=None)
            txt_h_pivot.value  = "Cliquez sur la courbe pour sélectionner le point de départ du contact"
            txt_h_pivot.color  = ft.Colors.GREY_600
            txt_h_pivot.update()
        except Exception as ex:
            txt_h_pivot.value = f"Erreur de lecture : {ex}"
            txt_h_pivot.color = ft.Colors.RED_700
            txt_h_pivot.update()

    # état zoom : None = vue complète, sinon (xmin, xmax, ymin, ymax)
    etat_zoom = {'xlim': None, 'ylim': None}

    def _dessiner_pivot(h_sel):
        ax_pivot.clear()
        h_arr = etat_pivot['h_exp']
        F_arr = etat_pivot['F_exp']
        if h_arr is None:
            ax_pivot.set_title("Aucune courbe chargée", color='gray', fontsize=10)
            chart_pivot.update()
            return
        ax_pivot.plot(h_arr, F_arr, color='steelblue', linewidth=1.4, label='F-h exp.')
        if h_sel is not None:
            idx = int(np.argmin(np.abs(h_arr - h_sel)))
            ax_pivot.axvline(h_sel, color='crimson', linestyle='--', linewidth=1.6, alpha=0.85)
            ax_pivot.plot(h_arr[idx], F_arr[idx], 'ro', markersize=9, zorder=5,
                          label=f'Origine h={h_sel:.3f}')
            ax_pivot.set_title(f"Origine sélectionnée : h = {h_sel:.4f} µm",
                               color='crimson', fontsize=10, fontweight='bold')
        else:
            ax_pivot.set_title("Cliquez sur le point de départ du contact (h = 0)",
                               color='gray', fontsize=10)
        ax_pivot.set_xlabel("h (µm)", fontsize=9)
        ax_pivot.set_ylabel("F (N)", fontsize=9)
        ax_pivot.legend(fontsize=9)
        ax_pivot.grid(True)
        if etat_zoom['xlim'] is not None:
            ax_pivot.set_xlim(etat_zoom['xlim'])
        if etat_zoom['ylim'] is not None:
            ax_pivot.set_ylim(etat_zoom['ylim'])
        chart_pivot.update()

    def _ylim_pour_xlim(xmin, xmax):
        """Y max adapté aux données visibles, Y min légèrement négatif pour voir le 0."""
        h_arr = etat_pivot['h_exp']
        F_arr = etat_pivot['F_exp']
        mask = (h_arr >= xmin) & (h_arr <= xmax)
        if not np.any(mask):
            return None
        F_vis = F_arr[mask]
        f_max = float(F_vis.max())
        return (-f_max * 0.20, f_max * 1.08)

    def _zoom(facteur: float):
        """facteur < 1 = zoom avant, > 1 = zoom arrière."""
        if etat_pivot['h_exp'] is None:
            return
        if etat_zoom['xlim'] is not None:
            xmin, xmax = etat_zoom['xlim']
        else:
            xmin, xmax = ax_pivot.get_xlim()
        # Si un point pivot est posé, le mettre au centre ; sinon garder le centre courant
        if etat_pivot['h_pivot'] is not None:
            centre = float(etat_pivot['h_pivot'])
        else:
            centre = (xmin + xmax) / 2.0
        demi   = (xmax - xmin) / 2.0 * facteur
        h_min  = float(etat_pivot['h_exp'][0])
        h_max  = float(etat_pivot['h_exp'][-1])
        new_xmin = max(centre - demi, h_min)
        new_xmax = min(centre + demi, h_max)
        etat_zoom['xlim'] = (new_xmin, new_xmax)
        etat_zoom['ylim'] = _ylim_pour_xlim(new_xmin, new_xmax)
        _dessiner_pivot(etat_pivot['h_pivot'])

    def _pan(direction: float):
        """Décale la vue de 20 % (direction : -1 gauche, +1 droite)."""
        if etat_pivot['h_exp'] is None:
            return
        h_data_min = float(etat_pivot['h_exp'][0])
        h_data_max = float(etat_pivot['h_exp'][-1])
        data_range = h_data_max - h_data_min
        # Lire depuis etat_zoom (source de vérité)
        if etat_zoom['xlim'] is not None:
            xmin, xmax = etat_zoom['xlim']
        else:
            xmin, xmax = h_data_min, h_data_max
        # Plafonner la largeur à la plage de données pour éviter le double-clamp
        width    = min(xmax - xmin, data_range)
        pas      = width * 0.20 * direction
        new_xmin = xmin + pas
        # Clamp unique : dériver new_xmax depuis new_xmin après clamp
        if new_xmin < h_data_min:
            new_xmin = h_data_min
        elif new_xmin + width > h_data_max:
            new_xmin = h_data_max - width
        new_xmax = new_xmin + width
        etat_zoom['xlim'] = (new_xmin, new_xmax)
        # Ne pas toucher au ylim en mode pan — évite le zoom/dézoom apparent
        # dû à la variation de F le long de la courbe
        _dessiner_pivot(etat_pivot['h_pivot'])

    def _zoom_reset():
        etat_zoom['xlim'] = None
        etat_zoom['ylim'] = None
        _dessiner_pivot(etat_pivot['h_pivot'])

    def _on_pivot_click(e):
        if etat_pivot['h_exp'] is None:
            txt_h_pivot.value = "⚠ Cliquez d'abord sur 'Charger la courbe'"
            txt_h_pivot.color = ft.Colors.ORANGE_700
            txt_h_pivot.update()
            return

        rendered_w = max(page.width - _RAIL_W - _PAD, 400)
        frac_x = e.local_x / rendered_w
        bbox   = ax_pivot.get_position()
        if not (bbox.x0 <= frac_x <= bbox.x0 + bbox.width):
            return

        t     = (frac_x - bbox.x0) / bbox.width
        xlim  = ax_pivot.get_xlim()
        h_val = xlim[0] + t * (xlim[1] - xlim[0])

        # Pas de snapping — valeur continue pour une sélection précise
        h_arr = etat_pivot['h_exp']
        h_sel = float(np.clip(h_val, h_arr[0], h_arr[-1]))
        # Premier point de données qui sera conservé (h >= h_sel)
        idx_premier = int(np.searchsorted(h_arr, h_sel))

        etat_pivot['h_pivot'] = h_sel
        _dessiner_pivot(h_sel)

        txt_h_pivot.value = f"✓  h_pivot = {h_sel:.4f} µm  ({len(h_arr) - idx_premier} points conservés)"
        txt_h_pivot.color = ft.Colors.GREEN_700
        txt_h_pivot.update()

    dest_pivot_label = ft.Text("Pivot Zéro", size=12)

    def _on_cb_change(e):
        input_decalage.visible    = cb_libre.value
        input_F_pivot.visible     = cb_faible.value
        cb_h_max_manuel.visible   = cb_fmax.value
        input_h_max.visible       = cb_fmax.value and cb_h_max_manuel.value
        dest_pivot_label.color    = None if cb_zero.value else ft.Colors.GREY_400
        dest_pivot_label.update()
        input_decalage.update()
        input_F_pivot.update()
        cb_h_max_manuel.update()
        input_h_max.update()

    cb_fmax.on_change   = _on_cb_change
    cb_libre.on_change  = _on_cb_change
    cb_faible.on_change = _on_cb_change
    cb_zero.on_change   = _on_cb_change
    cb_h_max_manuel.on_change = _on_cb_change

    # ── files ─────────────────────────────────────────────────────────────
    txt_bdd    = ft.Text("Aucun fichier sélectionné", italic=True, size=11)
    txt_courbe = ft.Text("Aucun fichier sélectionné", italic=True, size=11)
    txt_profil = ft.Text("Aucun fichier sélectionné", italic=True, size=11)
    txt_statut = ft.Text("", color=ft.Colors.BLUE_700, weight=ft.FontWeight.BOLD)

    # ── log ───────────────────────────────────────────────────────────────
    log_box = ft.TextField(
        multiline=True, read_only=True,
        min_lines=1,
        value="", text_size=11,
        bgcolor=ft.Colors.GREY_100,
        border_color=ft.Colors.BLACK12,
        expand=True,
    )

    def _log_append(msg: str):
        log_box.value = (log_box.value or "") + msg + "\n"
        log_box.update()

    def _log_clear():
        log_box.value = ""
        log_box.update()

    # ── live chart ────────────────────────────────────────────────────────
    fig_live, (ax_resid, ax_traj) = plt.subplots(1, 2, figsize=(9, 3.8))
    fig_live.subplots_adjust(wspace=0.38, bottom=0.18, left=0.09, right=0.97)
    ax_resid.set_title("Résidu (log)", color="gray", fontsize=9)
    ax_traj.set_title("Trajectoire (σ_y, n)", color="gray", fontsize=9)
    ax_resid.grid(True); ax_traj.grid(True)
    chart_live = MatplotlibChart(fig_live, expand=True)

    history_resid, history_x, history_y = [], [], []

    def _reset_live():
        history_resid.clear(); history_x.clear(); history_y.clear()
        ax_resid.clear(); ax_traj.clear()
        ax_resid.set_title("Résidu (log)", color="gray", fontsize=9)
        ax_traj.set_title("Trajectoire (σ_y, n)", color="gray", fontsize=9)
        ax_resid.grid(True); ax_traj.grid(True)
        chart_live.update()

    def _update_live(params_cb, iteration, resid):
        history_resid.append(abs(resid) if resid != 0 else 1e-30)
        history_x.append(params_cb['x'].value)
        history_y.append(params_cb['y'].value)
        if len(history_resid) % 10 != 0:
            return
        ax_resid.clear()
        ax_resid.plot(history_resid, color='tab:blue', linewidth=1)
        ax_resid.set_yscale('log')
        ax_resid.set_title("Résidu (log)", fontsize=9)
        ax_resid.grid(True)
        ax_traj.clear()
        ax_traj.plot(history_x, history_y, '-o', color='tab:blue', markersize=3, linewidth=1)
        if history_x:
            ax_traj.plot(history_x[-1], history_y[-1], 'ro', markersize=7)
        ax_traj.set_xlabel("σ_y", fontsize=8)
        ax_traj.set_ylabel("n",   fontsize=8)
        ax_traj.set_title("Trajectoire (σ_y, n)", fontsize=9)
        ax_traj.grid(True)
        chart_live.update()

    # ── result chart ──────────────────────────────────────────────────────
    fig_res, ax_res = plt.subplots(figsize=(6, 4.2))
    fig_res.subplots_adjust(left=0.13, bottom=0.14, right=0.97, top=0.9)
    ax_res.set_xlabel("Déformation ε", fontsize=9)
    ax_res.set_ylabel("Contrainte σ (MPa)", fontsize=9)
    ax_res.set_title("Loi d'écrouissage de Hollomon", fontsize=10)
    ax_res.grid(True)
    chart_res = MatplotlibChart(fig_res, expand=True)

    def _courbe_hollomon(sig_y, n, E):
        eps_y = sig_y / E
        def_i, con_i = [], []
        for d in np.linspace(0, eps_y, 10):
            def_i.append(d); con_i.append(E * d)
        d = eps_y
        while d < 0.45:
            def_i.append(d)
            con_i.append(loi_hollomon(sig_y, n, d, E=E))
            d *= 1.05
        return def_i, con_i

    def _afficher_resultats(resultats: dict, E_val: float):
        ax_res.clear()
        noms_labels = [METHODES_STYLE[k][3] for k in resultats]
        ax_res.set_title("Loi d'écrouissage — " + " / ".join(noms_labels), fontsize=9)
        ax_res.set_xlabel("Déformation ε", fontsize=9)
        ax_res.set_ylabel("Contrainte σ (MPa)", fontsize=9)
        for nom, (df, opt) in resultats.items():
            couleur, style, marker, label = METHODES_STYLE[nom]
            def_i, con_i = _courbe_hollomon(opt['sigma_y'], opt['n'], E_val)
            ax_res.plot(def_i, con_i, color=couleur, linewidth=2, linestyle=style,
                        label=f"{label}  σ_y={opt['sigma_y']:.1f}  n={opt['n']:.4f}")
            ax_res.scatter(df['epsilon'], df['sigma'], color=couleur, s=35,
                           zorder=5, marker=marker)
        ax_res.legend(fontsize=7)
        ax_res.grid(True)
        chart_res.update()

    # ── export ────────────────────────────────────────────────────────────
    _resultats_courants: dict = {}

    def _exporter_excel(chemin):
        import io
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        try:
            resultats = _resultats_courants.get('resultats')
            if not resultats:
                return
            E_val = _resultats_courants.get('E', 210000.0)
            path = chemin if chemin.endswith('.xlsx') else chemin + '.xlsx'

            wb = Workbook()

            # ── Feuille Résumé ──────────────────────────────────────────
            ws_sum = wb.active
            ws_sum.title = "Résumé"
            entetes = ["Méthode", "σ_y (MPa)", "n", "Rayon num (µm)"]
            for col, h in enumerate(entetes, 1):
                c = ws_sum.cell(row=1, column=col, value=h)
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="4472C4")
                c.font = Font(bold=True, color="FFFFFF")
                c.alignment = Alignment(horizontal="center")
            for row, (nom, (df, opt)) in enumerate(resultats.items(), 2):
                label = METHODES_STYLE[nom][3]
                rayon = opt.get('rayon_num')
                ws_sum.cell(row=row, column=1, value=label)
                ws_sum.cell(row=row, column=2, value=round(opt['sigma_y'], 3))
                ws_sum.cell(row=row, column=3, value=round(opt['n'], 6))
                ws_sum.cell(row=row, column=4, value=round(rayon, 4) if rayon else "—")
            for col in range(1, 5):
                ws_sum.column_dimensions[get_column_letter(col)].width = 20

            # ── Graphe Hollomon global (toutes méthodes) ─────────────────
            fig_xl, ax_xl = plt.subplots(figsize=(7, 4.5))
            for nom, (df, opt) in resultats.items():
                couleur, style, marker, label = METHODES_STYLE[nom]
                def_i, con_i = _courbe_hollomon(opt['sigma_y'], opt['n'], E_val)
                ax_xl.plot(def_i, con_i, color=couleur, linewidth=2, linestyle=style,
                           label=f"{label}  σ_y={opt['sigma_y']:.1f}  n={opt['n']:.4f}")
                ax_xl.scatter(df['epsilon'], df['sigma'], color=couleur, s=40,
                              zorder=5, marker=marker)
            ax_xl.set_xlabel("Déformation ε")
            ax_xl.set_ylabel("Contrainte σ (MPa)")
            ax_xl.set_title("Loi d'écrouissage de Hollomon")
            ax_xl.legend(fontsize=8)
            ax_xl.grid(True)
            buf = io.BytesIO()
            fig_xl.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig_xl)
            buf.seek(0)
            img = XLImage(buf)
            img.anchor = "F2"
            ws_sum.add_image(img)

            # ── Une feuille par méthode ──────────────────────────────────
            for nom, (df, opt) in resultats.items():
                label = METHODES_STYLE[nom][3]
                couleur = METHODES_STYLE[nom][0]
                ws = wb.create_sheet(title=label[:31])

                # En-têtes données
                cols_df = list(df.columns)
                for col, h in enumerate(cols_df, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor="D9E1F2")
                    c.alignment = Alignment(horizontal="center")

                for row_i, row_data in enumerate(df.itertuples(index=False), 2):
                    for col_i, val in enumerate(row_data, 1):
                        ws.cell(row=row_i, column=col_i, value=round(float(val), 6))

                for col in range(1, len(cols_df) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 16

                # σ_y et n finaux
                last_row = len(df) + 3
                ws.cell(row=last_row, column=1, value="σ_y final (MPa)").font = Font(bold=True)
                ws.cell(row=last_row, column=2, value=round(opt['sigma_y'], 3))
                ws.cell(row=last_row + 1, column=1, value="n final").font = Font(bold=True)
                ws.cell(row=last_row + 1, column=2, value=round(opt['n'], 6))

                # Graphe Hollomon par méthode
                fig_m, ax_m = plt.subplots(figsize=(6, 4))
                def_i, con_i = _courbe_hollomon(opt['sigma_y'], opt['n'], E_val)
                ax_m.plot(def_i, con_i, color=couleur, linewidth=2,
                          label=f"σ_y={opt['sigma_y']:.1f}  n={opt['n']:.4f}")
                ax_m.scatter(df['epsilon'], df['sigma'], color=couleur, s=50, zorder=5)
                ax_m.set_xlabel("Déformation ε")
                ax_m.set_ylabel("Contrainte σ (MPa)")
                ax_m.set_title(f"Hollomon — {label}")
                ax_m.legend(fontsize=8)
                ax_m.grid(True)
                buf_m = io.BytesIO()
                fig_m.savefig(buf_m, format='png', dpi=150, bbox_inches='tight')
                plt.close(fig_m)
                buf_m.seek(0)
                img_m = XLImage(buf_m)
                img_m.anchor = "H2"
                ws.add_image(img_m)

                # Graphe trajectoire d'optimisation (sous le tableau)
                buf_live = _resultats_courants.get(f'live_{nom}')
                if buf_live:
                    buf_live.seek(0)
                    img_live = XLImage(buf_live)
                    img_live.anchor = f"A{last_row + 4}"
                    ws.add_image(img_live)

            wb.save(path)
            page.overlay.append(
                ft.SnackBar(ft.Text(f"Export : {path}"),
                            bgcolor=ft.Colors.GREEN_700, open=True))
            page.update()
        except Exception as ex:
            import traceback
            txt_statut.value = f"Erreur export : {ex}"
            txt_statut.update()
            traceback.print_exc()

    # ── buttons ───────────────────────────────────────────────────────────
    progress_bar = ft.ProgressBar(visible=False)
    btn_run  = ft.ElevatedButton(
        "Lancer", icon=ft.Icons.PLAY_ARROW,
        bgcolor=ft.Colors.BLUE_700, color=ft.Colors.WHITE,
    )
    btn_stop = ft.ElevatedButton(
        "Stopper", icon=ft.Icons.STOP,
        bgcolor=ft.Colors.RED_100, color=ft.Colors.RED_900,
        visible=False,
    )
    btn_save = ft.OutlinedButton("Exporter Excel", icon=ft.Icons.DOWNLOAD, visible=False)

    def action_stopper(e):
        etat["stop"] = True
        btn_stop.text = "Arrêt en cours…"
        btn_stop.disabled = True
        btn_stop.update()

    btn_stop.on_click = action_stopper

    def _nom_fichier_auto():
        courbe = os.path.splitext(os.path.basename(txt_courbe.value))[0][:18].replace(" ", "_")
        methodes = [m for m, cb in [('fm', cb_fmax), ('li', cb_libre), ('fa', cb_faible), ('z', cb_zero), ('fr', cb_frelatif)] if cb.value]
        suffix = "-".join(methodes) if methodes else "opt"
        return f"{courbe}_{suffix}.xlsx"

    def _action_export(e):
        def _ouvrir():
            try:
                res = subprocess.run(
                    ["osascript", "-e", 'POSIX path of (choose folder with prompt "Choisir le dossier de sauvegarde")'],
                    capture_output=True, text=True, timeout=60
                )
                folder = res.stdout.strip()
                if folder:
                    _exporter_excel(os.path.join(folder, _nom_fichier_auto()))
            except Exception:
                pass
        threading.Thread(target=_ouvrir, daemon=True).start()

    btn_save.on_click = _action_export

    def _ouvrir_dans_preview():
        try:
            tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tmp.close()
            fig_res.savefig(tmp.name, dpi=200, bbox_inches='tight')
            subprocess.Popen(["open", tmp.name])
        except Exception as ex:
            txt_statut.value = f"Erreur agrandissement : {ex}"
            txt_statut.update()

    # ── navigation ────────────────────────────────────────────────────────
    view_config  = ft.Container(expand=True, visible=True,  padding=20)
    view_pivot   = ft.Container(expand=True, visible=False, padding=20)
    view_results = ft.Container(expand=True, visible=False, padding=20)

    def _navigate(idx: int):
        rail.selected_index  = idx
        view_config.visible  = (idx == IDX_CONFIG)
        view_pivot.visible   = (idx == IDX_PIVOT)
        view_results.visible = (idx == IDX_RESULTS)
        page.update()

    def _on_nav_change(e):
        _navigate(e.control.selected_index)

    rail = ft.NavigationRail(
        selected_index=IDX_CONFIG,
        label_type=ft.NavigationRailLabelType.ALL,
        min_width=90,
        group_alignment=-1.0,
        on_change=_on_nav_change,
        destinations=[
            ft.NavigationRailDestination(
                icon=ft.Icons.SETTINGS_OUTLINED,
                selected_icon=ft.Icons.SETTINGS,
                label="Config",
            ),
            ft.NavigationRailDestination(
                icon=ft.Icons.TOUCH_APP_OUTLINED,
                selected_icon=ft.Icons.TOUCH_APP,
                label_content=dest_pivot_label,
            ),
            ft.NavigationRailDestination(
                icon=ft.Icons.BAR_CHART_OUTLINED,
                selected_icon=ft.Icons.BAR_CHART,
                label="Résultats",
            ),
        ],
    )

    # ── compute thread ────────────────────────────────────────────────────
    def lancer(e):
        if etat["en_cours"]:
            return

        methodes_actives = []
        if cb_fmax.value:     methodes_actives.append('fmax')
        if cb_libre.value:    methodes_actives.append('libre')
        if cb_faible.value:   methodes_actives.append('faible')
        if cb_zero.value:     methodes_actives.append('zero')
        if cb_frelatif.value: methodes_actives.append('frelatif')

        if not methodes_actives:
            page.overlay.append(
                ft.SnackBar(ft.Text("⚠ Cochez au moins une méthode."),
                            bgcolor=ft.Colors.ORANGE_700, open=True))
            page.update()
            return

        if 'zero' in methodes_actives and etat_pivot['h_pivot'] is None:
            page.overlay.append(
                ft.SnackBar(
                    ft.Text("⚠ Pivot Zéro : sélectionnez d'abord l'origine (vue Pivot Zéro)."),
                    bgcolor=ft.Colors.ORANGE_700, open=True))
            page.update()
            return

        if "Aucun" in txt_bdd.value or "Aucun" in txt_courbe.value:
            page.overlay.append(
                ft.SnackBar(ft.Text("⚠ Chargez la BDD (.py) et la courbe (.dat) d'abord."),
                            bgcolor=ft.Colors.ORANGE_700, open=True))
            page.update()
            return

        try:
            E_val     = float(input_E.value)
            sig_guess = float(input_sig_y.value)
            n_guess   = float(input_n.value)
        except ValueError:
            page.overlay.append(
                ft.SnackBar(ft.Text("⚠ Paramètres numériques invalides."),
                            bgcolor=ft.Colors.RED_700, open=True))
            page.update()
            return

        try:
            decalage_init = float(input_decalage.value.replace(',', '.'))
        except ValueError:
            decalage_init = 0.0

        try:
            F_pivot_val = float(input_F_pivot.value.replace(',', '.'))
        except ValueError:
            F_pivot_val = 10.0

        h_max_val = None
        if cb_h_max_manuel.value and input_h_max.value.strip():
            try:
                h_max_val = float(input_h_max.value.replace(',', '.'))
            except ValueError:
                h_max_val = None

        etat["en_cours"] = True
        etat["stop"]     = False
        _reset_live()
        _log_clear()
        _resultats_courants.clear()

        btn_run.visible      = False
        btn_stop.visible     = True
        btn_stop.disabled    = False
        btn_stop.text        = "Stopper"
        btn_save.visible     = False
        progress_bar.visible = True
        txt_statut.value     = "Calcul en cours…"
        _navigate(IDX_RESULTS)
        page.update()

        portions = np.linspace(1.0, 0.3, 15)

        def thread_calcul():
            chemin_bdd    = txt_bdd.value
            chemin_courbe = txt_courbe.value
            try:
                bdd_module = charger_bdd_module(chemin_bdd)
                kwargs_communs = dict(
                    E=E_val,
                    sig_y_guess=sig_guess,
                    n_guess=n_guess,
                    chemin_courbe=chemin_courbe,
                    bdd_module=bdd_module,
                    portions=portions,
                    chemins_maillage=CHEMINS_MAILLAGE,
                    callback=_update_live,
                    stop_flag=lambda: etat["stop"],
                    log_callback=_log_append,
                )

                resultats_methodes = {}

                for nom in methodes_actives:
                    if etat["stop"]:
                        raise _AU_fmax()

                    _log_append(f"\n{'═'*38}\n  {METHODES_STYLE[nom][3]}\n{'═'*38}")
                    txt_statut.value = f"{METHODES_STYLE[nom][3]} : calcul en cours…"
                    txt_statut.update()
                    _reset_live()

                    if nom == 'fmax':
                        df, opt = executer_optimisation_fmax(
                            **kwargs_communs, h_max_override=h_max_val)
                    elif nom == 'libre':
                        df, opt = executer_optimisation_libre(
                            **kwargs_communs, decalage_initial=decalage_init)
                    elif nom == 'faible':
                        df, opt = executer_optimisation_faible(
                            **kwargs_communs, F_pivot=F_pivot_val)
                    elif nom == 'zero':
                        df, opt = executer_optimisation_zero(
                            **kwargs_communs, h_pivot=etat_pivot['h_pivot'])
                    elif nom == 'frelatif':
                        df, opt = executer_optimisation_frelatif(**kwargs_communs)

                    resultats_methodes[nom] = (df, opt)
                    import io as _io
                    _buf_live = _io.BytesIO()
                    fig_live.savefig(_buf_live, format='png', dpi=130, bbox_inches='tight')
                    _buf_live.seek(0)
                    _resultats_courants[f'live_{nom}'] = _buf_live
                    _log_append(f"  ✓ σ_y = {opt['sigma_y']:.2f} MPa  |  n = {opt['n']:.4f}")

                _resultats_courants['resultats'] = resultats_methodes
                _resultats_courants['E'] = E_val

                parties = []
                for nom, (df, opt) in resultats_methodes.items():
                    parties.append(
                        f"{METHODES_STYLE[nom][3]}: σ_y={opt['sigma_y']:.2f}  n={opt['n']:.4f}"
                    )
                r_exp = float(input_a.value.replace(',', '.'))
                if r_exp > 0:
                    r_num = resultats_methodes[premier][1].get('rayon_num')
                    if r_num:
                        err = abs(r_num - r_exp) / r_exp * 100
                        parties.append(f"Rayon={r_num:.2f} µm (écart {err:.1f}%)")
                txt_statut.value = "  |  ".join(parties)

                _afficher_resultats(resultats_methodes, E_val)
                btn_save.visible = True

            except _ARRET_CLASSES:
                txt_statut.value = "⚠ Calcul stoppé par l'utilisateur."
            except Exception as ex:
                import traceback
                txt_statut.value = f"Erreur : {ex}"
                traceback.print_exc()
            finally:
                etat["en_cours"]     = False
                progress_bar.visible = False
                btn_run.visible      = True
                btn_stop.visible     = False
                txt_statut.update()
                page.update()

        threading.Thread(target=thread_calcul, daemon=True).start()

    btn_run.on_click = lancer

    # ── view CONFIG ───────────────────────────────────────────────────────
    input_E.width     = 200
    input_a.width     = 200
    input_sig_y.width = 200
    input_n.width     = 200

    panneau_methodes = ft.Container(
        padding=16,
        border=ft.border.all(1, ft.Colors.BLACK12),
        border_radius=10,
        content=ft.Column([
            ft.Text("Méthodes", weight=ft.FontWeight.BOLD, size=13),
            ft.Row([cb_fmax, cb_libre, cb_faible, cb_zero, cb_frelatif], spacing=10),
            ft.Row([input_decalage, input_F_pivot, cb_h_max_manuel, input_h_max], spacing=10),
        ], spacing=10),
    )

    panneau_params = ft.Container(
        padding=16,
        border=ft.border.all(1, ft.Colors.BLACK12),
        border_radius=10,
        content=ft.Column([
            ft.Text("Paramètres", weight=ft.FontWeight.BOLD, size=13),
            ft.Row([input_E, input_a], spacing=10),
            ft.Row([input_sig_y, input_n], spacing=10),
        ], spacing=10),
    )

    def _ligne_fichier(label, txt_widget, choisir_fn):
        txt_widget.size   = 11
        txt_widget.italic = True
        txt_widget.no_wrap = True
        txt_widget.overflow = ft.TextOverflow.ELLIPSIS
        txt_widget.max_lines = 1
        return ft.Row([
            ft.ElevatedButton(label, icon=ft.Icons.FOLDER_OPEN,
                              on_click=lambda _: choisir_fn(),
                              style=ft.ButtonStyle(padding=ft.padding.symmetric(horizontal=10))),
            ft.Container(content=txt_widget, expand=True),
        ], spacing=8)

    panneau_fichiers = ft.Container(
        padding=16,
        border=ft.border.all(1, ft.Colors.BLACK12),
        border_radius=10,
        expand=True,
        content=ft.Column([
            ft.Text("Fichiers de données", weight=ft.FontWeight.BOLD, size=13),
            _ligne_fichier("BDD (.py)",               txt_bdd,    lambda: _choisir_fichier(txt_bdd)),
            _ligne_fichier("Courbe F-h (.dat)",        txt_courbe, lambda: _choisir_fichier(txt_courbe)),
            _ligne_fichier("Profil indenteur (.dat)",  txt_profil, lambda: _choisir_fichier(txt_profil)),
        ], spacing=10),
    )

    view_config.content = ft.Column(
        expand=True,
        spacing=14,
        controls=[
            ft.Text("Configuration",
                    style=ft.TextThemeStyle.HEADLINE_SMALL,
                    weight=ft.FontWeight.BOLD),
            ft.Divider(height=1),
            ft.Row([panneau_methodes, panneau_params, panneau_fichiers],
                   spacing=14, vertical_alignment=ft.CrossAxisAlignment.START),
            ft.Row(
                [btn_run, btn_stop, btn_save, progress_bar],
                alignment=ft.MainAxisAlignment.START,
                spacing=14,
            ),
            txt_statut,
            ft.Container(
                padding=ft.padding.only(left=10, right=10, bottom=8, top=6),
                border=ft.border.all(1, ft.Colors.BLACK12),
                border_radius=10,
                expand=True,
                content=ft.Column([
                    ft.Text("Journal de calcul", size=10,
                            color=ft.Colors.GREY_600, weight=ft.FontWeight.W_500),
                    log_box,
                ], spacing=4, expand=True),
            ),
        ],
    )

    # ── view PIVOT ZERO ───────────────────────────────────────────────────
    view_pivot.content = ft.Column(
        expand=True,
        spacing=14,
        controls=[
            ft.Row([
                ft.Icon(ft.Icons.TOUCH_APP, color=ft.Colors.RED_700, size=22),
                ft.Text("Pivot Zéro — Sélection de l'origine",
                        style=ft.TextThemeStyle.HEADLINE_SMALL,
                        weight=ft.FontWeight.BOLD,
                        color=ft.Colors.RED_800),
            ], spacing=8),
            ft.Divider(height=1),
            ft.Text(
                "Cliquez sur 'Charger la courbe' pour afficher la courbe F-h expérimentale, "
                "puis cliquez directement sur le graphe pour définir le point de départ du contact "
                "(ce point deviendra h = 0 dans l'optimisation).",
                size=12, color=ft.Colors.GREY_700,
            ),
            ft.Row([
                ft.ElevatedButton(
                    "Charger la courbe F-h",
                    icon=ft.Icons.SHOW_CHART,
                    on_click=lambda _: _charger_courbe_pivot(),
                    bgcolor=ft.Colors.RED_50,
                    color=ft.Colors.RED_900,
                ),
                txt_h_pivot,
            ], spacing=14),
            ft.Row([
                ft.Text("Zoom :", size=11, color=ft.Colors.GREY_700),
                ft.IconButton(
                    icon=ft.Icons.CHEVRON_LEFT,
                    tooltip="Déplacer à gauche",
                    icon_size=20,
                    on_click=lambda _: _pan(-1),
                ),
                ft.IconButton(
                    icon=ft.Icons.ADD_CIRCLE_OUTLINE,
                    tooltip="Zoom avant (×2)",
                    icon_size=20,
                    on_click=lambda _: _zoom(0.5),
                ),
                ft.IconButton(
                    icon=ft.Icons.REMOVE_CIRCLE_OUTLINE,
                    tooltip="Zoom arrière (÷2)",
                    icon_size=20,
                    on_click=lambda _: _zoom(2.0),
                ),
                ft.IconButton(
                    icon=ft.Icons.CHEVRON_RIGHT,
                    tooltip="Déplacer à droite",
                    icon_size=20,
                    on_click=lambda _: _pan(+1),
                ),
                ft.IconButton(
                    icon=ft.Icons.FIT_SCREEN,
                    tooltip="Réinitialiser la vue",
                    icon_size=20,
                    on_click=lambda _: _zoom_reset(),
                ),
            ], spacing=2),
            ft.Stack(
                expand=True,
                height=_PIVOT_H_PX,
                controls=[
                    ft.Container(
                        content=chart_pivot,
                        expand=True,
                        height=_PIVOT_H_PX,
                    ),
                    ft.GestureDetector(
                        content=ft.Container(
                            expand=True,
                            height=_PIVOT_H_PX,
                            bgcolor=ft.Colors.TRANSPARENT,
                        ),
                        mouse_cursor=ft.MouseCursor.PRECISE,
                        on_tap_down=_on_pivot_click,
                    ),
                ],
            ),
            ft.Text(
                "Astuce : choisissez le premier point de contact visible sur la courbe. "
                "Tous les points avant cette valeur de h seront ignorés dans l'optimisation.",
                size=10, color=ft.Colors.GREY_500, italic=True,
            ),
        ],
    )

    # ── view RESULTS ──────────────────────────────────────────────────────
    view_results.content = ft.Column(
        expand=True,
        spacing=14,
        controls=[
            ft.Text("Résultats",
                    style=ft.TextThemeStyle.HEADLINE_SMALL,
                    weight=ft.FontWeight.BOLD),
            ft.Divider(height=1),
            txt_statut,
            ft.Row(
                expand=True,
                vertical_alignment=ft.CrossAxisAlignment.START,
                controls=[
                    ft.Container(
                        expand=True,
                        border=ft.border.all(1, ft.Colors.BLACK12),
                        border_radius=10,
                        padding=10,
                        content=ft.Column([
                            ft.Text("Suivi de l'optimisation (live)",
                                    size=11, color=ft.Colors.GREY_600),
                            chart_live,
                        ], expand=True, spacing=4),
                    ),
                    ft.Container(
                        expand=True,
                        border=ft.border.all(1, ft.Colors.BLACK12),
                        border_radius=10,
                        padding=10,
                        content=ft.Column([
                            ft.Row([
                                ft.Text("Loi de Hollomon",
                                        size=11, color=ft.Colors.GREY_600),
                                ft.IconButton(
                                    icon=ft.Icons.OPEN_IN_FULL,
                                    tooltip="Agrandir dans Preview",
                                    icon_size=16,
                                    on_click=lambda _: _ouvrir_dans_preview(),
                                ),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            chart_res,
                        ], expand=True, spacing=4),
                    ),
                ],
            ),
            ft.Row([btn_save], alignment=ft.MainAxisAlignment.END),
        ],
    )

    # ── page layout ───────────────────────────────────────────────────────
    page.add(
        ft.Row(
            expand=True,
            spacing=0,
            controls=[
                rail,
                ft.VerticalDivider(width=1),
                ft.Container(
                    expand=True,
                    content=ft.Stack(
                        expand=True,
                        controls=[view_config, view_pivot, view_results],
                    ),
                ),
            ],
        )
    )
